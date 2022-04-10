from sklearn import svm
import openpyxl
from pathlib import Path
import numpy as np


def eigen_vectors(matrix):
    w, v = np.linalg.eig(matrix)
    lambda1 = 0
    for i in range(len(w)):
        if w[i] > w[lambda1]:
            lambda1 = i
    lambda2 = len(w) - lambda1 - 1
    for i in range(len(w)):
        if w[i] > w[lambda2] and i != lambda1:
            lambda2 = i
    result = [v[:, lambda1], v[:, lambda2]]
    return result


def proj(u, v):
    x = np.dot(u, v)
    y = np.dot(v, v)
    result = (x / y) * v
    return result


class PCA:

    def __init__(self, address):
        self.features = []
        xlsx_file = Path(address)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        self.sheet = wb_obj.active
        self.extract_features(self.sheet)
        self.data = []

    def extract_features(self, sheet):
        feature = []
        for column in sheet.iter_cols(max_col=4):
            for cell in column[1:]:
                feature.append(cell.value)
            self.features.append(feature)
            feature = []
        self.features = np.reshape(self.features, (4, self.sheet.max_row - 1))
        return

    def covariance_matrix(self):
        matrix = []
        row = []
        for i in range(4):
            for j in range(4):
                row.append(np.cov(self.features[i], self.features[j])[0][1])
            matrix.append(row)
            row = []
        return matrix

    def execute(self):
        evs = eigen_vectors(self.covariance_matrix())
        v1 = np.array(evs[0])
        v2 = np.array(evs[1])
        for i in range(4):
            self.features[i] = self.features[i] - np.mean(self.features[i])
        for i in range(self.sheet.max_row - 1):
            data = self.features[:, i]
            projection = proj(data, v1) + proj(data, v2)
            self.data.append(projection)
        return


if __name__ == "__main__":
    pca = PCA("dataset2.xlsx")
    l = pca.sheet["A"]
    data = []
    for i in range(pca.sheet.max_row - 1):
        data.append(pca.features[:, i])
    clf = svm.SVC(kernel="linear", gamma="auto")
    clf.fit(data, pca.data)

    pca.execute()
    clf.fit(data, pca.data)
